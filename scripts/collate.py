#!/usr/bin/env python3
"""
collate.py
==========

Bundle every CSV produced by ``pipeline.py`` into a single Excel workbook,
``outputs/results.xlsx``, with one sheet per logical results section.

Usage
-----
    python scripts/collate.py
    python scripts/collate.py --out outputs/my_results.xlsx

The first sheet (``README``) is an index that describes every other sheet
and the source CSV it was loaded from. Sheets that depend on a missing CSV
are silently skipped — re-run ``pipeline.py`` first to populate them.
"""

from __future__ import annotations

import argparse
import sys
from datetime import datetime
from pathlib import Path
from typing import List, Optional, Tuple

# Windows: force UTF-8 stdout so the bullet/check characters in our prints
# don't trip up the default cp1252 codec.
try:
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")
    sys.stderr.reconfigure(encoding="utf-8", errors="replace")
except Exception:
    pass

import numpy as np
import pandas as pd

PROJECT_ROOT = Path(__file__).resolve().parents[1]
OUT_DIR = PROJECT_ROOT / "outputs"


# Each entry: (sheet_name, source_csv, description)
# Excel sheet names are capped at 31 characters.
SHEETS: List[Tuple[str, str, str]] = [
    ("01_main_tda_epochs",          "tda_epoch_features_all.csv",
     "Main TDA: per-epoch H0/H1 persistence on EEG Fpz-Cz across W/N1/N2/N3/REM. "
     "Up to 30 epochs per stage per night, m=10, τ=2, seed=0."),
    ("02_main_tda_stage_summary",   "tda_stage_summary_all.csv",
     "Mean/SD/N for each H0/H1 metric grouped by sleep stage (collapsed across nights)."),

    ("03_robustness_grid_epochs",   "tda_robustness_grid_epochs.csv",
     "Per-epoch H1 features computed across the m∈{6,8,10,12} × τ∈{1,2,4} embedding grid. "
     "Up to 25 epochs per stage per night."),
    ("04_robust_mixedlm_omnibus",   "tda_robustness_mixedlm_omnibus.csv",
     "Likelihood-ratio omnibus tests for K0 (within-subject z of H1 metrics) "
     "fitted by mixed-LM (Powell, REML=False) per (channel, m, τ)."),
    ("05_robust_planned_contrasts", "tda_robustness_mixedlm_planned_contrasts.csv",
     "Planned contrasts (REM−W, REM−N3, N1−N3) with Holm-corrected p-values, per (channel, m, τ)."),

    ("06_baseline_epochs",          "baseline_epoch_features_all.csv",
     "Per-epoch baseline EEG features: log band power (δ/θ/α/σ/β), spectral entropy, "
     "permutation entropy, LZ76 complexity."),
    ("07_baseline_mixedlm_omnibus", "baseline_mixedlm_omnibus.csv",
     "Baseline LR omnibus tests (mixed-LM Powell, REML=False)."),
    ("08_baseline_planned",         "baseline_mixedlm_planned_contrasts.csv",
     "Baseline planned contrasts (REM−W, REM−N3, N1−N3) with Holm correction."),

    ("09_wake_subclass_epochs",     "wake_epoch_subclasses.csv",
     "Track-2 per-epoch wake QC features and labels (W_quiet / W_active_ocular / W_bad). "
     "Recording-percentile thresholds + 6-vote rule."),
    ("10_wake_subclass_summary",    "wake_subclass_summary.csv",
     "Per-night counts and percentages for each wake subclass."),
    ("11_wake_qc_epoch_table",      "wake_qc_epoch_table.csv",
     "Track-3 richer per-epoch wake-QC table with per-subject MAD-based thresholds."),

    ("12_tda_wake_subclasses",      "tda_epoch_features_wake_subclasses.csv",
     "Track-2 TDA features on REM + W_quiet + W_active_ocular epochs."),
    ("13_tda_wake_stage_summary",   "tda_stage_summary_wake_subclasses.csv",
     "Stage summary for the track-2 wake-subclass TDA features."),
    ("14_wake_mixedlm_omnibus",     "tda_wake_subclasses_mixedlm_omnibus.csv",
     "Track-2 wake-subclass mixed-LM omnibus tests on K0_tot/K0_max/K0_cnt."),
    ("15_wake_planned_contrasts",   "tda_wake_subclasses_mixedlm_planned_contrasts.csv",
     "Track-2 planned contrasts: REM−W_quiet, REM−W_active_ocular, W_quiet−W_active_ocular."),

    ("16_corrected_epochs_manifest","corrected_epochs_manifest.csv",
     "Manifest of EOG-regressed EEG NPZ bundles, with regression β per recording."),

    ("17_track3_raw_omnibus",       "raw_wake_mixedlm_omnibus.csv",
     "Track-3 raw EEG mixed-LM omnibus (lbfgs), wake-subclass stages."),
    ("18_track3_raw_contrasts",     "raw_wake_mixedlm_planned_contrasts.csv",
     "Track-3 raw EEG planned contrasts (Holm-corrected)."),
    ("19_track3_corr_omnibus",      "corrected_wake_mixedlm_omnibus.csv",
     "Track-3 EOG-corrected EEG mixed-LM omnibus."),
    ("20_track3_corr_contrasts",    "corrected_wake_mixedlm_planned_contrasts.csv",
     "Track-3 EOG-corrected EEG planned contrasts."),
    ("21_track3_eog_omnibus",       "eog_wake_mixedlm_omnibus.csv",
     "Track-3 EOG-channel mixed-LM omnibus (negative control)."),
    ("22_track3_eog_contrasts",     "eog_wake_mixedlm_planned_contrasts.csv",
     "Track-3 EOG-channel planned contrasts."),

    ("23_wake_robust_grid",         "wake_subclass_robustness_grid.csv",
     "Wake-subclass robustness grid: H1 features for m∈{8,10,12} × τ∈{1,2,3}."),
    ("24_wake_robust_omnibus",      "wake_subclass_robustness_mixedlm_omnibus.csv",
     "Wake-subclass robustness mixed-LM omnibus across the grid."),
    ("25_wake_robust_contrasts",    "wake_subclass_robustness_planned_contrasts.csv",
     "Wake-subclass robustness planned contrasts across the grid."),

    ("26_increm_model_fits",        "incremental_k0_vs_bandpower_model_fits.csv",
     "Binomial GLM fits (subject fixed effects) for the K0_tot vs band-power incremental analysis."),
    ("27_increm_lr_tests",          "incremental_k0_vs_bandpower_lr_tests.csv",
     "Likelihood-ratio tests: A↔B (add K0 to band power) and C↔D (add band power to K0)."),
    ("28_increm_coefficients",      "incremental_k0_vs_bandpower_coefficients.csv",
     "All non-subject coefficients across all four models (β, OR, 95% CI)."),

    ("29_review_wake_means",        "review_wake_subclass_stage_means.csv",
     "Stage-level means / SDs / N (subject-aggregated) for K0_tot/K0_max/K0_cnt across wake subclasses."),
    ("30_review_wake_dz",           "review_wake_subclass_paired_effect_sizes.csv",
     "Paired Cohen's d_z for each wake-subclass contrast."),
    ("31_review_wake_summary",      "review_wake_subclass_summary_table.csv",
     "Combined wake-subclass summary: omnibus + contrast + effect size on one row per metric/contrast."),

    ("32_review_base_means",        "review_baseline_wake_subclass_stage_means.csv",
     "Baseline metric stage means across wake subclasses."),
    ("33_review_base_dz",           "review_baseline_wake_subclass_paired_effect_sizes.csv",
     "Baseline Cohen's d_z for each wake-subclass contrast."),
    ("34_review_base_summary",      "review_baseline_wake_subclass_summary_table.csv",
     "Combined baseline wake-subclass summary table."),

    ("35_review_increm_summary",    "review_incremental_k0_vs_bandpower_summary.csv",
     "Incremental analysis: AIC/BIC/LR/p/β/OR for both REM-vs-wake-subclass contrasts, all four models."),
    ("36_review_increm_k0_only",    "review_incremental_k0_vs_bandpower_k0_only.csv",
     "Incremental review focused on whether K0 improves a band-power-only baseline (Model A → Model B)."),
    ("37_review_increm_bandpower",  "review_incremental_k0_vs_bandpower_bandpower_terms.csv",
     "Term-level coefficients (K0_tot + 4 band-power terms) from Model B."),

    ("38_compare_REM_vs_wake_wide", "comparison_table_rem_vs_wake_subclasses.csv",
     "Headline manuscript-ready comparison table: K0 vs every baseline metric for REM vs wake subclasses."),
    ("39_compare_REM_vs_wake_long", "comparison_table_rem_vs_wake_subclasses_long.csv",
     "Long-format version of the comparison table (one row per metric × contrast × source)."),

    ("40_supplementary_table",      "supplementary_table_wake_subclass_and_incremental_results.csv",
     "Supplementary table: wake-subclass contrasts + incremental K0 beyond band power."),
    ("41_supplementary_long",       "supplementary_table_wake_subclass_and_incremental_results_long.csv",
     "Long-format supplementary table."),

    # ── Revision-round additions ────────────────────────────────────────────
    ("42_stage_descriptives",       "stage_descriptives_all.csv",
     "Per-stage descriptives of K0_tot (mean, SD, median, IQR, n_subjects) for W/N1/N2/N3/REM."),
    ("43_all_pairwise_contrasts",   "stage_all_pairwise_contrasts.csv",
     "All 10 pairwise stage contrasts on K0_tot with Holm-corrected p-values."),
    ("44_stage_monotonicity",       "stage_monotonicity.csv",
     "Per-subject Spearman correlation of K0 across stage rank (W → REM); one-sample t-test against zero."),
    ("45_cohort_replication",       "cohort_replication_contrasts.csv",
     "Headline planned contrasts refit on Sleep-Cassette vs Sleep-Telemetry separately, "
     "for both K0 and every baseline metric."),
    ("46_subsampling_stability",    "subsampling_stability_contrasts.csv",
     "Post-hoc subsampling at caps {5,10,15,20,25,30} × 10 replicates; one row per "
     "(cap, replicate, contrast) showing contrast-estimate stability."),
    ("47_bootstrap_contrasts",      "bootstrap_contrasts.csv",
     "1000-iteration subject-level bootstrap of REM−W, REM−N3, N1−N3 with percentile 95% CIs."),
    ("48_embedding_ami",            "embedding_ami.csv",
     "Per-recording average mutual information AMI(τ) for τ ∈ 1..20 on a 30-recording subset."),
    ("49_embedding_fnn",            "embedding_fnn.csv",
     "Per-recording false-nearest-neighbours fraction at m ∈ 1..15 (τ = 2) on a 30-recording subset."),
    ("50_embedding_summary",        "embedding_diagnostics_summary.csv",
     "AMI/FNN-derived suggestions for τ and m; compares to the τ=2, m=10 used in the manuscript."),
    ("51_pz_oz_epochs",             "tda_epoch_features_pz_oz.csv",
     "Per-epoch TDA features computed on EEG Pz-Oz (where present) for the multi-channel control."),
    ("52_pz_oz_omnibus",            "tda_pz_oz_mixedlm_omnibus.csv",
     "Pz-Oz mixed-LM omnibus test for the K0_tot × stage effect."),
    ("53_pz_oz_contrasts",          "tda_pz_oz_mixedlm_planned_contrasts.csv",
     "Pz-Oz planned contrasts (REM−W, REM−N3, N1−N3) with Holm-corrected p-values."),
    ("54_preproc_sensitivity",      "preprocessing_sensitivity_contrasts.csv",
     "Headline contrasts across bandpass {0.5-30, 0.5-40, 0.5-45, 1-40} × sfreq {50,100,128} on 30 random nights."),
    ("55_statistical_diagnostics",  "statistical_diagnostics.csv",
     "Shapiro-Wilk normality test on residuals, Levene's homoscedasticity across stages, "
     "skewness and excess kurtosis."),
    ("56_classification_loso",      "classification_loso_metrics.csv",
     "Per-fold LOSO classification metrics (AUC, balanced accuracy, F1, sensitivity, specificity)."),
    ("57_classification_summary",   "classification_summary.csv",
     "LOSO classification summary: mean ± SD across folds for each (target, feature_set, model). "
     "Targets: REM-vs-W, REM-vs-NREM, REM-vs-other. Feature sets: K0_only, bandpower_only, combined."),
]


def _autosize(ws, df: pd.DataFrame, max_width: int = 60):
    """Best-effort column auto-sizing for openpyxl worksheets."""
    try:
        from openpyxl.utils import get_column_letter
    except Exception:
        return
    for i, col in enumerate(df.columns, start=1):
        try:
            series = df[col].astype(str)
            max_len = series.str.len().max()
            # NaN-safe: an entirely-NaN column can yield NaN here on some
            # pandas / Python combinations (Python 3.12 in particular), and
            # int(NaN) raises "cannot convert float NaN to integer".
            if pd.isna(max_len):
                max_len = 0
            width = min(max_width, max(len(str(col)) + 2, int(max_len) + 2))
            ws.column_dimensions[get_column_letter(i)].width = width
        except Exception:
            # Auto-sizing is best-effort; never let it abort a sheet write.
            continue


def _write_readme(writer, available: List[Tuple[str, str, str, int]], n_missing: int):
    rows = [{
        "Sheet": s, "Source CSV": src, "Rows": n,
        "Description": desc,
    } for (s, src, desc, n) in available]
    df = pd.DataFrame(rows)
    df.to_excel(writer, sheet_name="README", index=False)
    ws = writer.sheets["README"]

    # Bold header
    try:
        from openpyxl.styles import Font, Alignment
        for cell in ws[1]:
            cell.font = Font(bold=True)
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = Alignment(wrap_text=True, vertical="top")
    except Exception:
        pass
    _autosize(ws, df, max_width=80)

    # Header banner row
    try:
        from openpyxl.styles import Font, PatternFill, Alignment
        ws.insert_rows(1, amount=3)
        ws["A1"] = "Topological Recurrence in EEG Dynamics — Consolidated Results"
        ws["A2"] = f"Built {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        ws["A3"] = (
            f"{len(available)} sheets present"
            + (f", {n_missing} sheets skipped (run pipeline.py to populate)" if n_missing else "")
        )
        ws["A1"].font = Font(bold=True, size=14)
        ws["A2"].font = Font(italic=True)
        ws.merge_cells("A1:D1"); ws.merge_cells("A2:D2"); ws.merge_cells("A3:D3")
    except Exception:
        pass


def main(argv=None):
    p = argparse.ArgumentParser(description="Collate every pipeline output into one Excel workbook.")
    p.add_argument("--out", type=str, default=str(OUT_DIR / "results.xlsx"),
                   help="Output xlsx path (default outputs/results.xlsx).")
    p.add_argument("--max-rows", type=int, default=200_000,
                   help="Truncate sheets larger than this to keep the file readable.")
    args = p.parse_args(argv)

    out_path = Path(args.out)
    if not out_path.is_absolute():
        out_path = (PROJECT_ROOT / out_path).resolve()
    out_path.parent.mkdir(parents=True, exist_ok=True)

    available: List[Tuple[str, str, str, int]] = []
    missing:   List[Tuple[str, str]] = []

    print(f"Building {out_path.relative_to(PROJECT_ROOT)} ...")
    try:
        writer = pd.ExcelWriter(out_path, engine="openpyxl")
    except ImportError:
        raise SystemExit(
            "openpyxl is required. Install it with:\n  pip install openpyxl"
        )

    with writer:
        # Placeholder for README — rewritten at end
        pd.DataFrame({"_": ["pending"]}).to_excel(writer, sheet_name="README", index=False)

        for sheet, src, desc in SHEETS:
            csv_path = OUT_DIR / src
            if not csv_path.exists():
                missing.append((sheet, src))
                print(f"  · skip {sheet:34s} (missing {src})")
                continue
            try:
                df = pd.read_csv(csv_path)
            except Exception as ex:
                print(f"  ! could not read {src}: {type(ex).__name__}: {ex}")
                missing.append((sheet, src))
                continue
            n = len(df)
            if n > args.max_rows:
                df = df.head(args.max_rows)
                print(f"  · truncated {src} from {n:,} to {args.max_rows:,} rows")
            try:
                df.to_excel(writer, sheet_name=sheet, index=False)
                _autosize(writer.sheets[sheet], df)
                available.append((sheet, src, desc, n))
                print(f"  ✓ {sheet:34s} ({n:,} rows)")
            except Exception as ex:
                print(f"  ! failed to write {sheet}: {type(ex).__name__}: {ex}")
                missing.append((sheet, src))

        # Re-write README on top
        del writer.sheets["README"]
        if "README" in writer.book.sheetnames:
            std = writer.book["README"]
            writer.book.remove(std)
        _write_readme(writer, available, n_missing=len(missing))
        # move README to first position
        try:
            idx = writer.book.sheetnames.index("README")
            if idx != 0:
                writer.book.move_sheet("README", offset=-idx)
        except Exception:
            pass

    print(f"\nDone. Wrote {len(available)} sheets to {out_path}")
    if missing:
        print(f"Skipped (missing source CSVs): {len(missing)}")
        for s, src_name in missing:
            print(f"  - {s}: {src_name}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
